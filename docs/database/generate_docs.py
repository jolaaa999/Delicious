"""
生成「人间烟火 (Delicious)」数据库文档 Excel
格式仿照 垂直调度数据表.xlsx
"""
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 样式定义 ──────────────────────────────────────────────
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

title_font = Font(name='微软雅黑', size=14, bold=True)
subtitle_font = Font(name='微软雅黑', size=12, bold=True)
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=10)
note_font = Font(name='微软雅黑', size=9, color='666666')

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subtitle_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
light_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

COL_WIDTHS = [22, 18, 40, 42, 48]


def setup_sheet(ws, title, tables):
    """tables: [(table_cn, table_en, fields)]
    fields: [(字段名, 类型, 说明, 约束, 备注)]
    """
    row = 1
    # ── 模块标题（合并 A1:E2） ──
    ws.merge_cells('A1:E2')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    row = 3

    for ti, (table_cn, table_en, fields) in enumerate(tables):
        # 空行
        row += 2

        # ── 表名行（合并） ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=f'{table_cn}（{table_en}）')
        cell.font = subtitle_font
        cell.alignment = center_align
        cell.fill = subtitle_fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = subtitle_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        # ── 表头行 ──
        headers = ['字段名', '类型', '说明', '约束', '备注']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

        # ── 字段行 ──
        for fi, (name, dtype, desc, constraint, remark) in enumerate(fields):
            vals = [name, dtype, desc, constraint, remark]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = normal_font
                cell.alignment = left_align if ci >= 3 else center_align
                cell.border = thin_border
                if fi % 2 == 0:
                    cell.fill = light_fill
            ws.row_dimensions[row].height = 20
            row += 1

    # ── 列宽 ──
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ============================================================================
# Sheet 1: 用户管理
# ============================================================================
ws1 = wb.active
ws1.title = '用户管理'

tables_user = [
    ('用户信息', 'users', [
        ('id', 'BIGINT UNSIGNED', '主键', 'PRIMARY KEY, AUTO_INCREMENT', ''),
        ('username', 'VARCHAR(64)', '登录用户名', 'UNIQUE, NOT NULL', ''),
        ('email', 'VARCHAR(128)', '邮箱', 'UNIQUE, DEFAULT NULL', '可选，未来可用于找回密码'),
        ('password_hash', 'VARCHAR(255)', '密码哈希', 'NOT NULL', 'bcrypt 加密存储，API 不返回'),
        ('nickname', 'VARCHAR(64)', '显示昵称', "NOT NULL, DEFAULT ''", '个人使用场景可设为"家人"'),
        ('avatar_url', 'VARCHAR(512)', '头像 URL', 'DEFAULT NULL', '指向本地 /uploads 或 Blob 存储'),
        ('status', 'TINYINT', '账户状态', 'NOT NULL, DEFAULT 1', '1=正常，0=禁用（软锁定）'),
        ('created_at', 'DATETIME(3)', '创建时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3)', '毫秒精度'),
        ('updated_at', 'DATETIME(3)', '更新时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3) ON UPDATE CURRENT_TIMESTAMP(3)', '自动更新'),
        ('deleted_at', 'DATETIME(3)', '软删除时间', 'DEFAULT NULL, INDEX', 'GORM 软删除，非空表示已删除'),
    ]),
]

setup_sheet(ws1, '用户管理', tables_user)

# 加索引说明
row = ws1.max_row + 3
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws1.cell(row=row, column=1, value='索引说明')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws1.cell(row=row, column=c).fill = subtitle_fill
    ws1.cell(row=row, column=c).border = thin_border

indexes = [
    ('uk_users_username', 'UNIQUE', 'username', '保证用户名唯一'),
    ('uk_users_email', 'UNIQUE', 'email', '保证邮箱唯一'),
    ('idx_users_deleted_at', 'INDEX', 'deleted_at', '加速软删除查询过滤'),
]
row += 1
for ci, h in enumerate(['索引名', '类型', '字段', '说明'], 1):
    cell = ws1.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1
for fi, (name, itype, col, desc) in enumerate(indexes):
    for ci, v in enumerate([name, itype, col, desc], 1):
        cell = ws1.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws1.cell(row=row, column=5).value = ''  # empty 5th col
    ws1.cell(row=row, column=5).border = thin_border
    row += 1

# ============================================================================
# Sheet 2: 百科菜谱
# ============================================================================
ws2 = wb.create_sheet('百科菜谱')

tables_encyclopedia = [
    ('百科菜谱信息', 'encyclopedia_recipes', [
        ('id', 'BIGINT UNSIGNED', '主键', 'PRIMARY KEY, AUTO_INCREMENT', ''),
        ('name', 'VARCHAR(128)', '菜名', 'NOT NULL, INDEX', '如"红烧肉"、"Kung Pao Chicken"'),
        ('description', 'TEXT', '简介/描述', 'DEFAULT NULL', '菜品的简要介绍'),
        ('cover_image_url', 'VARCHAR(512)', '封面图 URL', 'DEFAULT NULL', '菜品展示图片'),
        ('category', 'VARCHAR(64)', '分类', 'DEFAULT NULL, INDEX', '如"家常菜"、"粤菜"、"川菜"、"西餐"等'),
        ('tags', 'JSON', '标签数组', 'DEFAULT NULL', '存储格式: ["快手菜","下饭菜","硬菜"]'),
        ('ingredients', 'JSON', '标准配料与重量', 'NOT NULL', '存储格式: [{"name":"五花肉","amount":500,"unit":"g","note":"切块"}]'),
        ('process_steps', 'JSON', '标准制作步骤', 'NOT NULL', '存储格式: [{"order":1,"content":"焯水...","duration_minutes":5}]'),
        ('source', 'VARCHAR(128)', '数据来源', 'DEFAULT NULL', '如"百科"、"spoonacular"、"themealdb"'),
        ('external_source', 'VARCHAR(32)', '外部来源标识', 'UNIQUE(组合), DEFAULT NULL', '如"spoonacular"、"themealdb"，用于去重'),
        ('external_id', 'VARCHAR(64)', '外部 ID', 'UNIQUE(组合), DEFAULT NULL', '在外部源中的唯一标识'),
        ('view_count', 'INT UNSIGNED', '浏览次数', 'NOT NULL, DEFAULT 0', '每次获取详情时自动 +1'),
        ('created_at', 'DATETIME(3)', '创建时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3)', ''),
        ('updated_at', 'DATETIME(3)', '更新时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3) ON UPDATE CURRENT_TIMESTAMP(3)', ''),
    ]),
]

setup_sheet(ws2, '百科菜谱', tables_encyclopedia)

row = ws2.max_row + 3
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws2.cell(row=row, column=1, value='索引说明')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws2.cell(row=row, column=c).fill = subtitle_fill
    ws2.cell(row=row, column=c).border = thin_border

indexes_ency = [
    ('idx_encyclopedia_name', 'INDEX', 'name', '加速按菜名搜索'),
    ('idx_encyclopedia_category', 'INDEX', 'category', '加速按分类筛选'),
    ('idx_encyclopedia_external', 'UNIQUE', '(external_source, external_id)', '防止外部数据重复导入'),
    ('ft_encyclopedia_name_desc', 'FULLTEXT', '(name, description)', 'MySQL 全文索引，支持中文分词搜索'),
]
row += 1
for ci, h in enumerate(['索引名', '类型', '字段', '说明'], 1):
    cell = ws2.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1
for fi, (name, itype, col, desc) in enumerate(indexes_ency):
    for ci, v in enumerate([name, itype, col, desc], 1):
        cell = ws2.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws2.cell(row=row, column=5).value = ''
    ws2.cell(row=row, column=5).border = thin_border
    row += 1

# JSON 字段结构说明
row += 2
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws2.cell(row=row, column=1, value='JSON 字段结构说明')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws2.cell(row=row, column=c).fill = subtitle_fill
    ws2.cell(row=row, column=c).border = thin_border
row += 1

# ingredients 子结构
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws2.cell(row=row, column=1, value='▶ ingredients / process_steps 子字段（Ingredient / ProcessStep 结构体）').font = Font(name='微软雅黑', size=10, bold=True)
row += 1
json_headers = ['子字段名', '类型', '说明', '必填', '备注']
for ci, h in enumerate(json_headers, 1):
    cell = ws2.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border
row += 1

ingredient_fields = [
    ('name', 'string', '食材名称', '是', '如"五花肉"、"生抽"'),
    ('amount', 'float64', '用量数值', '是', '如 500、2.5'),
    ('unit', 'string', '用量单位', '是', '如"g"、"勺"、"个"、"ml"'),
    ('note', 'string', '备注说明', '否', '如"切块"、"焯水后用"'),
]
for fi, (n, t, d, r, rm) in enumerate(ingredient_fields):
    for ci, v in enumerate([n, t, d, r, rm], 1):
        cell = ws2.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

row += 1
process_fields = [
    ('order', 'int', '步骤序号', '是', '从 1 开始递增'),
    ('content', 'string', '步骤描述文本', '是', '如"五花肉切块，冷水下锅焯水去腥"'),
    ('duration_minutes', 'int', '预计耗时（分钟）', '否', '如 45，为 null 表示不限制'),
    ('image_url', 'string', '步骤配图 URL', '否', '可为每个步骤单独配图'),
]
for fi, (n, t, d, r, rm) in enumerate(process_fields):
    for ci, v in enumerate([n, t, d, r, rm], 1):
        cell = ws2.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    row += 1


# ============================================================================
# Sheet 3: 我的菜谱
# ============================================================================
ws3 = wb.create_sheet('我的菜谱')

tables_my = [
    ('我的菜谱主表', 'my_recipes', [
        ('id', 'BIGINT UNSIGNED', '主键', 'PRIMARY KEY, AUTO_INCREMENT', ''),
        ('user_id', 'BIGINT UNSIGNED', '所属用户 ID', 'FOREIGN KEY → users.id, NOT NULL, INDEX', 'CASCADE 删除：删除用户时级联删除其菜谱'),
        ('name', 'VARCHAR(128)', '菜名', 'NOT NULL', '用户自定义菜名'),
        ('current_version_id', 'BIGINT UNSIGNED', '当前生效版本 ID', 'FOREIGN KEY → recipe_versions.id, DEFAULT NULL', '指向最新版本，新建时先为 NULL，创建版本后回填'),
        ('user_rating', 'TINYINT UNSIGNED', '个人评分 1-5 星', 'DEFAULT NULL, INDEX', '用户对成品的评分，NULL 表示未评分'),
        ('cover_image_url', 'VARCHAR(512)', '列表封面图', 'DEFAULT NULL', '通常取自当前版本的首张图片（冗余字段，加速列表展示）'),
        ('encyclopedia_recipe_id', 'BIGINT UNSIGNED', '关联百科菜谱 ID', 'FOREIGN KEY → encyclopedia_recipes.id, DEFAULT NULL, INDEX', '用于与基准版本对比，SET NULL 当百科菜谱被删除'),
        ('created_at', 'DATETIME(3)', '创建时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3), INDEX', ''),
        ('updated_at', 'DATETIME(3)', '更新时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3) ON UPDATE CURRENT_TIMESTAMP(3)', ''),
        ('deleted_at', 'DATETIME(3)', '软删除时间', 'DEFAULT NULL, INDEX', 'GORM 软删除'),
    ]),
]

setup_sheet(ws3, '我的菜谱', tables_my)

row = ws3.max_row + 3
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws3.cell(row=row, column=1, value='索引说明')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws3.cell(row=row, column=c).fill = subtitle_fill
    ws3.cell(row=row, column=c).border = thin_border

indexes_my = [
    ('idx_my_recipes_user_id', 'INDEX', 'user_id', '按用户筛选菜谱'),
    ('idx_my_recipes_user_rating', 'INDEX', 'user_rating', '按评分筛选/排序'),
    ('idx_my_recipes_created_at', 'INDEX', 'created_at', '按时间排序'),
    ('idx_my_recipes_deleted_at', 'INDEX', 'deleted_at', '软删除过滤'),
    ('idx_my_recipes_encyclopedia', 'INDEX', 'encyclopedia_recipe_id', '关联百科菜谱查询'),
]
row += 1
for ci, h in enumerate(['索引名', '类型', '字段', '说明'], 1):
    cell = ws3.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1
for fi, (name, itype, col, desc) in enumerate(indexes_my):
    for ci, v in enumerate([name, itype, col, desc], 1):
        cell = ws3.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws3.cell(row=row, column=5).value = ''
    ws3.cell(row=row, column=5).border = thin_border
    row += 1

# 版本控制流程
row += 2
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws3.cell(row=row, column=1, value='▶ 版本控制写入流程（应用层事务）')
cell.font = Font(name='微软雅黑', size=10, bold=True)
cell.fill = subtitle_fill
for c in range(1, 6):
    ws3.cell(row=row, column=c).fill = subtitle_fill
    ws3.cell(row=row, column=c).border = thin_border
row += 1
flow_steps = [
    '1. 新建菜谱: INSERT my_recipes (current_version_id = NULL)',
    '2. 创建首个版本: INSERT recipe_versions (version_number = 1)',
    '3. 回填版本指针: UPDATE my_recipes SET current_version_id = 新版本.id',
    '4. 编辑菜谱: INSERT recipe_versions (version_number = MAX(version_number) + 1)',
    '5. 更新指针: UPDATE my_recipes SET current_version_id = 新版本.id',
    '★ 核心原则: 编辑时仅 INSERT 新版本行，绝不 UPDATE 旧版本行（不可变记录）',
]
for step in flow_steps:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws3.cell(row=row, column=1, value=step).font = Font(name='微软雅黑', size=10)
    ws3.cell(row=row, column=1).alignment = left_align
    row += 1


# ============================================================================
# Sheet 4: 版本管理
# ============================================================================
ws4 = wb.create_sheet('版本管理')

tables_ver = [
    ('菜谱版本详情', 'recipe_versions', [
        ('id', 'BIGINT UNSIGNED', '主键', 'PRIMARY KEY, AUTO_INCREMENT', '每个版本一个唯一 ID'),
        ('recipe_id', 'BIGINT UNSIGNED', '所属菜谱 ID', 'FOREIGN KEY → my_recipes.id, NOT NULL, INDEX', 'CASCADE 删除：删除菜谱时级联删除所有版本'),
        ('version_number', 'INT UNSIGNED', '版本号', 'NOT NULL, UNIQUE(与 recipe_id 组合)', '从 1 递增，每个菜谱独立计数'),
        ('ingredients', 'JSON', '食材与重量', 'NOT NULL', '结构同百科菜谱，[{"name":"生抽","amount":2,"unit":"勺"}]'),
        ('process_steps', 'JSON', '制作步骤', 'NOT NULL', '结构同百科菜谱，[{"order":1,"content":"...","duration_minutes":null}]'),
        ('process_text', 'TEXT', '步骤纯文本', 'DEFAULT NULL', '可选，兼容简单文本录入（不推荐）'),
        ('images', 'JSON', '版本图片列表', 'DEFAULT NULL', '存储格式: ["https://cdn/1.jpg","https://cdn/2.jpg"]'),
        ('commit_msg', 'VARCHAR(255)', '修改备注', "NOT NULL, DEFAULT ''", '类似 Git commit message，如"减盐，调整火候"'),
        ('created_at', 'DATETIME(3)', '创建时间', 'NOT NULL, DEFAULT CURRENT_TIMESTAMP(3), INDEX', '版本创建时间（不可变）'),
    ]),
]

setup_sheet(ws4, '版本管理', tables_ver)

row = ws4.max_row + 3
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws4.cell(row=row, column=1, value='索引说明')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws4.cell(row=row, column=c).fill = subtitle_fill
    ws4.cell(row=row, column=c).border = thin_border

indexes_ver = [
    ('uk_recipe_version', 'UNIQUE', '(recipe_id, version_number)', '保证同一菜谱下版本号唯一'),
    ('idx_recipe_versions_recipe_id', 'INDEX', 'recipe_id', '加速按菜谱查询所有版本'),
    ('idx_recipe_versions_created_at', 'INDEX', 'created_at', '加速按时间排序'),
]
row += 1
for ci, h in enumerate(['索引名', '类型', '字段', '说明'], 1):
    cell = ws4.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1
for fi, (name, itype, col, desc) in enumerate(indexes_ver):
    for ci, v in enumerate([name, itype, col, desc], 1):
        cell = ws4.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws4.cell(row=row, column=5).value = ''
    ws4.cell(row=row, column=5).border = thin_border
    row += 1

# 版本对比算法说明
row += 2
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws4.cell(row=row, column=1, value='▶ 版本对比算法（diff 包）')
cell.font = Font(name='微软雅黑', size=10, bold=True)
cell.fill = subtitle_fill
for c in range(1, 6):
    ws4.cell(row=row, column=c).fill = subtitle_fill
    ws4.cell(row=row, column=c).border = thin_border
row += 1

diff_info = [
    '算法复杂度: O(n+m)，n 和 m 分别为两个版本的配料/步骤数量',
    '配料匹配: 按 name 归一化（忽略大小写、去空格）建立 HashMap 索引',
    '单位归一化: "g"↔"克"、"kg"↔"千克"、"ml"↔"毫升"、"l"↔"升" 视为等价',
    '浮点比较: 使用 1e-6 epsilon 公差比较用量',
    '差异类型: unchanged（不变）/ added（新增）/ removed（删除）/ modified（修改）',
    '用量增量: modified 类型附带 amount_delta（target.amount - base.amount）',
    '输出摘要: 自动生成中文摘要，如"新增 2 项配料，修改 1 个步骤"',
]
for info in diff_info:
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws4.cell(row=row, column=1, value=info).font = Font(name='微软雅黑', size=10)
    ws4.cell(row=row, column=1).alignment = left_align
    row += 1


# ============================================================================
# Sheet 5: 关系模型
# ============================================================================
ws5 = wb.create_sheet('关系模型')

# Title
ws5.merge_cells('A1:E2')
ws5['A1'] = '关系模型 — 表关系与外键依赖'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align
ws5.row_dimensions[1].height = 28
ws5.row_dimensions[2].height = 28

row = 5
# ── 关系表 ──
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws5.cell(row=row, column=1, value='表间关系一览')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws5.cell(row=row, column=c).fill = subtitle_fill
    ws5.cell(row=row, column=c).border = thin_border
row += 1

rel_headers = ['父表（被引用）', '子表（含外键）', '外键字段', '关系类型', '删除策略']
for ci, h in enumerate(rel_headers, 1):
    cell = ws5.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

relations = [
    ('users', 'my_recipes', 'user_id', '1 : N（一个用户可有多个菜谱）', 'CASCADE — 删除用户时级联删除其所有菜谱'),
    ('my_recipes', 'recipe_versions', 'recipe_id', '1 : N（一个菜谱可有多个版本）', 'CASCADE — 删除菜谱时级联删除所有版本'),
    ('recipe_versions', 'my_recipes', 'current_version_id', '1 : 0..1（菜谱指向其中一个版本）', 'SET NULL — 删除版本时清空菜谱的当前版本指针'),
    ('encyclopedia_recipes', 'my_recipes', 'encyclopedia_recipe_id', '1 : N（一个百科菜谱可被多个我的菜谱引用）', 'SET NULL — 删除百科菜谱时保留我的菜谱但断开引用'),
]
for ri, (parent, child, fk, rel_type, on_delete) in enumerate(relations):
    vals = [parent, child, fk, rel_type, on_delete]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if ri % 2 == 0:
            cell.fill = light_fill
    ws5.row_dimensions[row].height = 22
    row += 1

# ── ER 文字描述 ──
row += 2
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws5.cell(row=row, column=1, value='实体关系描述（ER 模型）')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws5.cell(row=row, column=c).fill = subtitle_fill
    ws5.cell(row=row, column=c).border = thin_border
row += 1

er_desc = [
    '',
    '┌──────────┐          ┌──────────────────────┐          ┌─────────────────┐',
    '│  users   │ 1 ─── N │     my_recipes       │ 1 ─── N │ recipe_versions │',
    '│          │◄────────│                      │◄────────│                 │',
    '│ id (PK)  │         │ id (PK)              │         │ id (PK)         │',
    '│ username │         │ user_id (FK)         │         │ recipe_id (FK)  │',
    '│ ...      │         │ current_version_id ──│────────►│ version_number  │',
    '└──────────┘         │ encyclopedia_recipe_id│        │ ingredients     │',
    '                     └──────────┬───────────┘         │ process_steps   │',
    '                                │                     │ commit_msg      │',
    '                                │ N                   │ ...             │',
    '                                │                     └─────────────────┘',
    '                     ┌──────────▼───────────┐',
    '                     │ encyclopedia_recipes │',
    '                     │                      │',
    '                     │ id (PK)              │',
    '                     │ name                 │',
    '                     │ ingredients (JSON)   │',
    '                     │ process_steps (JSON) │',
    '                     │ external_source/id   │',
    '                     │ ...                  │',
    '                     └──────────────────────┘',
    '',
]
for line in er_desc:
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws5.cell(row=row, column=1, value=line)
    c.font = Font(name='Consolas', size=10) if '─' in line or '│' in line or '┌' in line or '└' in line or '┘' in line or '┐' in line or '►' in line or '▼' in line else normal_font
    c.alignment = left_align
    row += 1

# ── 第三范式分析 ──
row += 2
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws5.cell(row=row, column=1, value='第三范式（3NF）合规性分析')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws5.cell(row=row, column=c).fill = subtitle_fill
    ws5.cell(row=row, column=c).border = thin_border
row += 1

nf3_headers = ['表名', '1NF（原子性）', '2NF（无部分依赖）', '3NF（无传递依赖）', '合规结论']
for ci, h in enumerate(nf3_headers, 1):
    cell = ws5.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

nf3_data = [
    ('users', '✅ 所有字段原子值', '✅ 单列主键，无复合主键，自动满足', '✅ 所有非键属性直接依赖于主键 id', '✅ 符合 3NF'),
    ('encyclopedia_recipes', '✅ 所有字段原子值; JSON 字段存储结构化数据，MySQL 中视为原子', '✅ 单列主键', '✅ 所有非键属性直接依赖于主键 id', '✅ 符合 3NF'),
    ('my_recipes', '✅ 所有字段原子值', '✅ 单列主键', '⚠ cover_image_url 从 current_version 的首张图片冗余而来（性能优化），严格来说存在传递依赖：id → current_version_id → images[0] → cover_image_url', '⚠ 近似 3NF（含一处有意的反范式冗余）'),
    ('recipe_versions', '✅ 所有字段原子值', '✅ 单列主键', '✅ 所有非键属性直接依赖于主键 id; version_number 与 recipe_id 组成唯一约束，不违反 3NF', '✅ 符合 3NF'),
]

for ri, (table, nf1, nf2, nf3, result) in enumerate(nf3_data):
    vals = [table, nf1, nf2, nf3, result]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(row=row, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        if ri % 2 == 0:
            cell.fill = light_fill
    ws5.row_dimensions[row].height = 36
    row += 1

# ── 数据流架构图 ──
row += 2
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws5.cell(row=row, column=1, value='核心数据流架构')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws5.cell(row=row, column=c).fill = subtitle_fill
    ws5.cell(row=row, column=c).border = thin_border
row += 1

dataflow = [
    '',
    '  ┌─────────────────────────────────────────────────────────────────┐',
    '  │                     数据流架构 (Data Flow)                       │',
    '  ├─────────────────────────────────────────────────────────────────┤',
    '  │                                                                 │',
    '  │  前端 (Vue3)  ──REST/JSON──►  Gin Router  ──►  Handler         │',
    '  │                                                        │         │',
    '  │                                                        ▼         │',
    '  │                                                    Service       │',
    '  │                                              ┌───────┴───────┐  │',
    '  │                                              │               │  │',
    '  │                                              ▼               ▼  │',
    '  │                                         Repository     外部 API  │',
    '  │                                        (数据库操作)   (联网搜索)  │',
    '  │                                              │     (翻译服务)    │',
    '  │                                              ▼                  │',
    '  │                                         GORM / PostgreSQL        │',
    '  │                                                                 │',
    '  │  联网搜索:                                                       │',
    '  │  关键词 → 中译英扩展 → Spoonacular API ─┐                        │',
    '  │                   └→ MealDB API ───────┤                        │',
    '  │                                         ▼                        │',
    '  │                              去重 → Upsert → encyclopedia_recipes│',
    '  │                                                                 │',
    '  │  翻译管道:                                                       │',
    '  │  encyclopedia_recipes ──► MyMemory API ──► 返回翻译后 JSON       │',
    '  │  (最大 6 并发 goroutine，分段翻译长文本)                         │',
    '  │                                                                 │',
    '  │  版本对比:                                                       │',
    '  │  my_recipes ──► recipe_versions ──► diff.Compare()              │',
    '  │  百科基准 ──► encyclopedia_recipes ──► diff.Compare()           │',
    '  │  O(n+m) HashMap 配料匹配 + 单位归一化                            │',
    '  │                                                                 │',
    '  └─────────────────────────────────────────────────────────────────┘',
    '',
]
for line in dataflow:
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws5.cell(row=row, column=1, value=line)
    c.font = Font(name='Consolas', size=10)
    c.alignment = left_align
    row += 1


# ============================================================================
# Sheet 6: 创新点与完善建议
# ============================================================================
ws6 = wb.create_sheet('创新与完善建议')

ws6.merge_cells('A1:E2')
ws6['A1'] = '数据操作创新点与完善建议'
ws6['A1'].font = title_font
ws6['A1'].alignment = center_align

row = 5
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws6.cell(row=row, column=1, value='一、已有创新点（软著亮点）')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws6.cell(row=row, column=c).fill = subtitle_fill
    ws6.cell(row=row, column=c).border = thin_border
row += 1

innovations = [
    ('版本控制系统（类 Git）',
     '菜谱修改采用 append-only 不可变版本模型，每次编辑 INSERT 新行而不 UPDATE 旧版本，'
     '通过 current_version_id 指针指向最新版本，支持版本历史追溯、版本对比和时间线。'
     '这是菜谱管理领域较少见的严谨版本控制设计。'),
    ('版本差异算法 (diff)',
     'O(n+m) 时间复杂度的配料/步骤对比算法，使用 HashMap 索引 + 名称归一化 + 单位归一化 '
     '（g↔克、kg↔千克、ml↔毫升），浮点数 epsilon 比较，自动生成中文摘要。'
     '支持版本间对比和与百科基准对比两种模式。'),
    ('多源聚合联网搜索',
     '关键词中译英扩展（内置词典 + 在线翻译 API）→ 多 Provider 策略模式 (Spoonacular + MealDB) '
     '→ 按 source:externalID 去重 → Upsert 写入百科库。支持在线数据自动缓存与过期刷新。'),
    ('并发翻译管道',
     'Semaphore 控制最大 6 并发 goroutine，分段翻译长文本绕过 API 长度限制，'
     '支持列表级（仅译菜名）和详情级（全字段翻译）两级粒度，对搜索结果和百科详情分别优化。'),
    ('JSON 自定义类型',
     '泛型 JSONSlice[T] 和 StringSlice 实现 GORM 的 Valuer/Scanner 接口，'
     '将复杂结构的序列化/反序列化封装在模型层，Service 层直接操作 Go 原生类型。'),
]
for title, desc in innovations:
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws6.cell(row=row, column=1, value=f'◆ {title}')
    c.font = Font(name='微软雅黑', size=10, bold=True)
    c.alignment = left_align
    row += 1
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws6.cell(row=row, column=1, value=f'   {desc}')
    c.font = normal_font
    c.alignment = left_align
    ws6.row_dimensions[row].height = 40
    row += 1

row += 2
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws6.cell(row=row, column=1, value='二、数据操作层面可创新的方向')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws6.cell(row=row, column=c).fill = subtitle_fill
    ws6.cell(row=row, column=c).border = thin_border
row += 1

future_innovations = [
    ('1. 食材归一化/同义词引擎',
     '建立 ingredient_aliases 表，将"鸡蛋/鸡子"、"生抽/酱油"等异名归一化到标准名称。'
     '在版本对比和搜索时自动进行同义词匹配，提升 diff 准确性和搜索召回率。'
     '可引入编辑距离（Levenshtein）做模糊匹配容错。'),
    ('2. 智能菜谱缩放（Scaling）',
     '在 recipe_versions 增加 servings 字段，基于线性插值算法实现跨份量的配料自动换算。'
     '考虑非线性调料（盐等）需要特殊处理的缩放规则表。'),
    ('3. 配料替换建议',
     '基于食材分类（蛋白质/蔬菜/调料）和营养特征，在对比时自动建议替代食材。'
     '例如：用户缺少"老抽"时建议用"生抽+糖"替代，并自动调整用量。'),
    ('4. 营养计算模块',
     '引入食材营养数据库（如 USDA），自动计算每个版本的卡路里、蛋白质、脂肪等营养数据，'
     '支持按份量换算，并输出营养标签。'),
    ('5. 菜谱推荐算法',
     '基于用户历史评分（user_rating）和食材偏好，使用协同过滤或基于内容的推荐算法，'
     '从百科菜谱库自动推荐用户可能感兴趣的菜谱。'),
    ('6. 购物清单自动生成',
     '根据选中的菜谱版本，聚合所有食材用量，按分类（蔬菜/肉类/调料）生成采购清单，'
     '支持多菜谱合并去重和按份量自动换算。'),
    ('7. 全文搜索增强',
     '为 my_recipes 和 recipe_versions 添加全文索引，支持按食材名称、步骤内容全文检索。'
     'PostgreSQL 可使用 tsvector + GIN 索引实现。'),
]
for title, desc in future_innovations:
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws6.cell(row=row, column=1, value=title)
    c.font = Font(name='微软雅黑', size=10, bold=True)
    c.alignment = left_align
    row += 1
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws6.cell(row=row, column=1, value=f'   {desc}')
    c.font = normal_font
    c.alignment = left_align
    ws6.row_dimensions[row].height = 50
    row += 1

row += 2
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws6.cell(row=row, column=1, value='三、当前需完善的地方')
cell.font = subtitle_font
cell.fill = subtitle_fill
for c in range(1, 6):
    ws6.cell(row=row, column=c).fill = subtitle_fill
    ws6.cell(row=row, column=c).border = thin_border
row += 1

improvements = [
    ('1. N+1 查询问题',
     'List() 方法中对每个 recipe 逐一调用 GetVersion() 获取当前版本号，存在 N+1 查询。'
     '建议：使用 JOIN 或批量查询（WHERE id IN (...)）一次性获取所有版本信息。', '性能'),
    ('2. 分类管理',
     'category 字段目前为自由文本，存在同义异名问题（如"家常菜"/"家常"）。'
     '建议：创建 categories 字典表，将 category 改为 FK 引用，实现标准化分类管理。', '数据一致性'),
    ('3. 标签管理',
     'tags 使用 JSON 数组存储，缺乏统一约束。建议：创建 tags 字典表和 recipe_tags 关联表，'
     '实现标签的标准化和多对多关联。', '数据一致性'),
    ('4. 删除确认机制',
     'my_recipes 使用软删除但缺少恢复机制。建议：增加回收站功能（30天可恢复）和真正的物理删除清理任务。', '功能完整性'),
    ('5. 上传文件清理',
     '版本更新后旧图片可能不再被引用，当前无垃圾回收机制。建议：增加图片引用计数或定期清理孤儿文件。', '存储管理'),
    ('6. 离线数据导出',
     '当前无数据导出功能。建议：支持 JSON/YAML 格式的菜谱数据导出/导入，便于数据迁移和备份。', '可移植性'),
    ('7. API 限流与缓存',
     '外部 API（翻译、菜谱搜索）无本地缓存层。建议：增加 Redis 缓存层，减少 API 调用次数和费用。', '性能/成本'),
]
improve_headers = ['问题', '描述与建议', '影响类别']
for ci, h in enumerate(improve_headers, 1):
    cell = ws6.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1
for fi, (title, desc, category) in enumerate(improvements):
    ws6.cell(row=row, column=1, value=title).font = Font(name='微软雅黑', size=10, bold=True)
    ws6.cell(row=row, column=2, value=desc).font = normal_font
    ws6.cell(row=row, column=3, value=category).font = normal_font
    for ci in range(1, 4):
        ws6.cell(row=row, column=ci).alignment = left_align
        ws6.cell(row=row, column=ci).border = thin_border
        if fi % 2 == 0:
            ws6.cell(row=row, column=ci).fill = light_fill
    # merge cells for description
    ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws6.row_dimensions[row].height = 36
    row += 1

# Column widths for sheet 6
for ci, w in enumerate([30, 80, 15], 1):
    ws6.column_dimensions[get_column_letter(ci)].width = w


# ============================================================================
# 保存
# ============================================================================
output_path = r'E:\PROJECT\Delicious\docs\database\人间烟火_数据库设计文档.xlsx'
wb.save(output_path)
print(f'Done: {output_path}')
